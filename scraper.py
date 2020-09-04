import requests
from bs4 import BeautifulSoup
import os


from lxml import html
import openpyxl
from openpyxl.utils import get_column_letter


def get_share_price(url):
    try:
        page = requests.get(url)
        html_soup = BeautifulSoup(page.text, 'html.parser')
        sp = html_soup.find('span', style = 'font-weight:bold').text
    except:
        sp = "error xD"
    return sp


class ExcelHandler():
    
    def __init__(self, path_excel):
        self.path_excel = path_excel
        self.__init_file()
        
    def __del__(self):
        self.__save_wb()
    
    def __save_wb(self):
        self.wb.save(filename=self.path_excel)
    
    def __init_sheet(self, title):
        if title in self.wb.sheetnames:
            return self.wb[title]
        else:
            return self.wb.create_sheet(title)
    
    def read_data(self, sheet_name):
        ws = self.__init_sheet(sheet_name)
        return ws.rows
    
    def clean_sheet(self, sheet_name):
        ws = self.__init_sheet(sheet_name)
        for row in ws.rows:
            for cell in row:
                cell.value = None
        self.__save_wb()
        
    def save_data(self, sheet_name, data, flag_autosize_columns=True, flag_overwrite=False):
        ws = self.__init_sheet(sheet_name)
        if flag_overwrite:
            for i, r in enumerate(data, start=1):
                for j, c in enumerate(r, start=1):
                    ws.cell(row=i, column=j).value = c
        else:
            for d in data:
                ws.append(d)
        if flag_autosize_columns:
            self.__autosize_columns(ws, data)
        self.__save_wb()
    
    def __init_file(self):
         if os.path.exists(self.path_excel):
             self.wb = openpyxl.load_workbook(filename=self.path_excel)
         else:
             self.wb = openpyxl.Workbook()
             self.__save_wb
    
    def __autosize_columns(self, ws, data):
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
                else:
                    column_widths += [len(cell)]
        
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width


def main_program(path_input):
    sheet_name = 'Sheet1'
    eh = ExcelHandler(path_input)
    data = eh.read_data(sheet_name)
    
    new_data = []
    for i, d in enumerate(data):
        if i == 0:
            headers = []
            for c in d:
                headers.append(c.value)
            new_data.append(headers)
        else:
            symbol = d[1].value
            url = f'https://stooq.pl/q/?s={symbol}'
            price = get_share_price(url)
            print('scraping', symbol)
            new_data.append([d[0].value, symbol, price, url])
    eh.save_data(sheet_name, new_data, True, True)
    print('done')
        

if __name__ == '__main__':
    path_input = 'Input.xlsx'
    main_program(path_input)



